from flask import Flask, render_template, url_for, request, send_file, session, redirect
from flask_sqlalchemy import SQLAlchemy
from io import BytesIO
application = Flask(__name__)
from zipfile import ZipFile
from PIL import Image
import os
import imagehash
from requests_html import HTMLSession, HTML
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import time
import io
application.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///application.db'
application.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
application.config['SQLALCHEMY_BINDS'] = {'db2': 'sqlite:///application.db2','db3': 'sqlite:///application.db3'}
from flask_mysqldb import MySQL
from werkzeug.utils import secure_filename
application.config['MYSQL_HOST'] = '127.0.0.1'
application.config['MYSQL_USER'] = 'root'
application.config['MYSQL_PASSWORD'] = 'xxxxx'
application.config['MYSQL_DB'] = 'testdb'


script_dir = os.path.dirname(__file__)
rel_path = "static"
savepath = os.path.join(script_dir, rel_path)

UPLOAD_FOLDER = savepath
application.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
mysql = MySQL(application)
application.secret_key = os.urandom(30).hex()




print(savepath)

@application.route('/')
@application.route('/homepage')
def homepage():
    session['datab'] = os.urandom(10).hex()
    return render_template('home.html')
    
@application.route('/uploadfiles', methods=['GET','POST'])

def index():

    if request.method == 'POST':
        for addedfile in request.files.getlist('filenamee'):
            file = addedfile
            ibes = session['datab']
            namefile = file.filename
            datafile = file.read()
            cursor = mysql.connection.cursor()
            cursor.execute(''' INSERT INTO testtable VALUES(%s,%s,%s)''',(ibes,namefile,datafile))
            mysql.connection.commit()
            cursor.close()
        return redirect(url_for("download"))


    else:
        return render_template('index.html')

   
@application.route('/uploadfiles/done', methods=['GET','POST'])  
def download():
    cursor = mysql.connection.cursor()
    cursor.execute("SELECT * FROM testtable")
    pictures = cursor.fetchall()
   # upload = Upload.query.filter_by(id=upload_id).first()
    # upload.filename = '33.jpg'
    # return send_file(BytesIO(upload.data), download_name=upload.filename, as_attachment=True)
    
    # print(newname)
    if request.method == 'POST':
        newname = request.form['lname']
        oldname = request.form['fname']
        return redirect(url_for("listallitems",lname=newname,fname=oldname))
    else:

            

    
        return render_template('form.html',pictures=pictures)
    
@application.route('/download/all<lname><fname>', methods=['GET'])
def listallitems(lname,fname):
    zipfilename = os.urandom(10).hex()
    memorryzip = f'{zipfilename}.zip'
    newname = lname
    oldname = fname
    cursor = mysql.connection.cursor()
    cursor.execute("SELECT * FROM testtable")
    pictures = cursor.fetchall()
    with ZipFile(memorryzip, 'w') as myzip:
        for c in pictures:
            print(c[0])
            with BytesIO() as tmp:
                tmp.write(c[2])
                tmp.seek(0)
                x = tmp.read()
                newnamee = c[1].replace(oldname,newname)
                
                myzip.writestr(newnamee,x)
    
    return send_file(f'{memorryzip}', as_attachment=True)


  
@application.route('/compareimages',methods=['GET','POST'])
def uploadcompare():
    
            
    if request.method == 'POST':
        for addedfile in request.files.getlist('filenamee'):
            file = addedfile
            ibes = session['datab']
            namefile = file.filename
            datafile = file.read()
            cursor = mysql.connection.cursor()
            cursor.execute(''' INSERT INTO comparetable VALUES(%s,%s,%s)''',(ibes,namefile,datafile))
            mysql.connection.commit()
            cursor.close()
        return redirect(url_for("compareimages"))
    else:
        return render_template('index.html')

@application.route('/compareimages/download',methods=['GET'])
def compareimages():

    listofhashes = []
    zipfilename = os.urandom(10).hex()
    memorryzip = f'{zipfilename}.zip'

    cursor = mysql.connection.cursor()
    cursor.execute("SELECT * FROM comparetable")
    picturestocompare = cursor.fetchall()
    with ZipFile(memorryzip, 'w') as myzip:
        for originalpicutre in picturestocompare:
            with BytesIO() as tmp:
                tmp.write(originalpicutre[2])
                tmp.seek(0)
                x = tmp.read()

                originalimagehash = imagehash.average_hash(Image.open(tmp))
                
                if originalimagehash in listofhashes:
                    print('dub')
                    newnamee = originalpicutre[1].replace(originalpicutre[1],f'dub{originalpicutre[1]}')

                else:
                    listofhashes.append(originalimagehash)
                    newnamee = originalpicutre[1]

                myzip.writestr(newnamee,x)
    
    return send_file(f'../{memorryzip}', as_attachment=True)

@application.route('/downloadexcel',methods=['GET','POST'])
def uploadexcel():
    ibes = os.urandom(3).hex()
    if request.method == 'POST':
        file = request.files['filenameee']
        filename = secure_filename(file.filename)
        file.save(os.path.join(application.config['UPLOAD_FOLDER'], filename))

        
        memorryzip = f'{ibes}.zip'
        savepath = os.path.join(UPLOAD_FOLDER,filename)
        print(savepath)
        deletelist = []
        with ZipFile(memorryzip, 'w') as myzip:

            with open (savepath,'rb') as ff:
                session = HTMLSession()
                count = 1
               
                wb = load_workbook(ff)
                ws = wb.active
                imagelink = ws['F'+str(count)].value
                print('opened workbook')
                column_f = ws['F']
                for cell in column_f:
                   
                    if imagelink == None or count > 10:
                        print('break')
                        break
                    if count >1:
                        if cell.value !=  ws['F'+str(count-1)].value and '.jpg' in cell.value and 'htt' in cell.value:
                            print(f'{cell}ffffffffffff')
                            time.sleep(1)
                            r2 = session.get(cell.value)
                            print(cell.value)
                            namefile = f'{count}.jpg'
                            datafile = r2.content
                            cursor = mysql.connection.cursor()
                            cursor.execute(''' INSERT INTO downloadedimages VALUES(%s,%s,%s,%s)''',(0,ibes,namefile,datafile))
                            mysql.connection.commit()
                            cursor.close()
                           
                            print('yes')

                                
                    count +=1
            cursor = mysql.connection.cursor()
            cursor.execute("SELECT * FROM downloadedimages")
            pictures = cursor.fetchall()
            for c in pictures:
                if c[1] == ibes:
                    print(c[0])
                    with BytesIO() as tmp:
                        tmp.write(c[3])
                        tmp.seek(0)
                        x = tmp.read()
                        # print(x)
                        # newname = 
                        newnamee = c[2]
                    
                        myzip.writestr(newnamee,x)


        os.remove(savepath)
        return send_file(f'../{memorryzip}', as_attachment=True)
    else:
        return render_template('excelupload.html')


if __name__ == "__main__":
    application.run(debug=True)