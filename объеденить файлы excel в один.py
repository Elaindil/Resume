import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from pathlib import Path
import glob
startTime = datetime.now()
txtfiles = []
for file in glob.glob("*.xlsm"):
    txtfiles.append(file)
print(txtfiles)

SOURCE_DIR = 'test'
excel_files = list(Path(SOURCE_DIR).glob("*.xlsm"))
print(excel_files)
wb = load_workbook('test.xlsx')

print('loaded workbook')
ws = wb['WorkSheet']



column_r = ws['R']
count2 = 1
count = 1



for file in txtfiles:
    print(file)
    wb2 = load_workbook(file)
    ws2 = wb2['result']
    column_e = ws2['E']
    count = 1
    print(count)
    for columne in column_e:

        columnb2 = ws['B'+str(count2)].value
        columnc2 = ws['C'+str(count2)].value
        columnd2 = ws['D'+str(count2)].value
        columne2 = ws['E'+str(count2)].value

        columna = ws2['A'+str(count)].value
        columnb = ws2['B'+str(count)].value
        columnc = ws2['C'+str(count)].value
        columnd = ws2['D'+str(count)].value
        if ws2['A'+str(count)].value == None:
            print('no')
            break 
        
        if count >1:
        
            count2 = count2 + 1
            ws['A'+str(count2)].value = columna
            ws['B'+str(count2)].value = columnb
            ws['c'+str(count2)].value = columnc
            ws['D'+str(count2)].value = columnd
            ws['E'+str(count2)].value = columne.value
            print(file)


        count = count + 1


wb.save('checked.xlsx')  
print(datetime.now() - startTime)   
