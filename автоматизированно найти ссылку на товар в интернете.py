from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

startTime = datetime.now()
options = Options()
options.add_argument("--headless")

driver = webdriver.Firefox(options=options)


wb = load_workbook('test.xlsx')
print('loaded workbook')
ws = wb['Sheet']

column_d = ws['D']

column_l = ws['L']

count = 1
linktocheck = ''

foundlinka = False



for columnd in column_d:
    if columnd.value == None:
        break  
    
    if count >1:
        if columnd.value !=  ws['D'+str(count-1)].value: #если новая ссылка то открывает её

            searchstring = columnd.value.replace(' ',' !')
            searchstring = searchstring.replace('цв.','цв.!')
            searchstring = searchstring.replace('!1','1')
            searchme = f"https://yandex.ru/search/?text={searchstring}"
            driver.get(searchme)
            print(searchme)

            continue_link = driver.find_elements(By.CLASS_NAME, 'Link.Link_theme_outer.Path-Item.link.path__item.link.organic__greenurl')
            for e in continue_link:
                e=e.get_attribute('href')
                if 'www.XXXX' in e:
                    foundlinka = True
                    foundlink = e

            
            time.sleep(3)
            linktocheck =foundlink
            print(linktocheck)
            print(count)

      
    
    if foundlinka == True:
        ws['R'+str(count)].value = linktocheck
    if count % 3000 == 0:
        wb.save(f'{count}addedlinks.xlsx') 

    count = count + 1
   
wb.save('addedlinks.xlsx')  
driver.close() 



print(datetime.now() - startTime)   

