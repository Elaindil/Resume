from PIL import Image
import imagehash

from pathlib import Path
import glob
import requests
from requests_html import HTMLSession, HTML
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
startTime = datetime.now()


wb = load_workbook('xxxx.xlsx')
print('loaded workbook')
ws = wb['WorkSheet']

column_a = ws['A']
session = HTMLSession()
column_r = ws['R']
ft = Font(color="FF0000")
count = 1
linktocheck = ''

foundlinka = False

count = 1
differencesfound = 0
for columna in column_a:
    
    if columna.value == None:
        print('stop')
        break 
    
    if count >1:
        if columna.value !=  ws['A'+str(count-1)].value: #если новая ссылка то открывает её
            linkcell = ws['R'+str(count)].value
            originalpic = ws['I'+str(count)].value
            origpiccompare= requests.get(originalpic, stream=True)
            
            r = session.get(linkcell)
            matches = r.html.find('img.gyoE4')
            ihash = imagehash.average_hash(Image.open(origpiccompare.raw))
            for match in matches:
                # time.sleep(1)
                newimage  = match.attrs["src"].replace('140.jpg','700.jpg')
                newimagecompare= requests.get(newimage, stream=True)
                ihash2 = imagehash.average_hash(Image.open(newimagecompare.raw))
                print(originalpic)
                print(newimage)
                print(ihash == ihash2)
                difference = ihash - ihash2
                if difference > 2:
                    print('error')
                    print(f'{count}')
                    colourme = ws['R'+str(count)]
                    colourme2 = ws['D'+str(count)]
                    colourme.font = ft
                    colourme2.font = ft
                      # hamming distance
                    differencesfound +=1
                    print(differencesfound)
                break


    count = count + 1


wb.save('xxx.xlsx')  
print(datetime.now() - startTime)  