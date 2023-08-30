from requests_html import HTMLSession, HTML
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import time
from datetime import datetime
import re



startTime = datetime.now()






wb = load_workbook('xxxx.xlsx')
print('Loaded new workbook') 
ws = wb['Sheet']

column_s = ws['R']

column_l = ws['L']

session = HTMLSession()

count = 0
neccessaryattributes = ['Торговая марка','Артикул','Страна производитель','Высота, см','В наборе, шт.','Серия','Размер упаковки (Длина ? Ш','Размер (Длина ? Ширина ? В','Вес брутто','Цвет','Можно мыть в посудомоечной машине','Диаметр, см','Вид упаковки','Материал','Размер упаковки (Длина × Ширина × Высота)','Вид упаковки','Назначение','Можно использовать в СВЧ-печи','Особенности','Объём, мл','Размер (Длина × Ширина × Высота)','Крышка','Индивидуальная упаковка','Размер, см','Толщина, мм','Размер (Длина × Ширина)','Толщина, мм','Размер, см','Количество персон','Вид крепежа','В наборе, шт.','Количество фотографий','Размер (Длина/Ширина)','Ширина, см','Высота, см','Выбор имени','Особенности','Дополнительный аксессуар','Составляющие набора','Тематика','Вид упаковки','Тематика праздника']
dictionary = {'Артикул производителя':'Артикул','Бренд':'Торговая марка','Вес коробки, в килограммах':'Вес брутто','Вес, в граммах':'Вес брутто','Длина коробки, в сантиметрах':'Размер упаковки (Длина × Ширина × Высота)','Материал чашки/кружки':'Материал','Материал':'Состав ткани','Высота изделия, в сантиметрах':'Высота, см','Особенности дизайна':'Тематика','Особенности дизайна':'Тематика праздника','Коллекция':'Серия','Страна-производитель':'Страна производитель','Форма':'Форма','Цвет':'Цвет','Объем, в миллилитрах':'Объём, мл','Тип тарелки':'Тип тарелки','Наименование модели':'Наименование модели','Высота коробки, в сантиметрах':'Размер упаковки (Длина × Ширина × Высота)','Высота, в сантиметрах':'Размер фото','Комплектация':'Составляющие набора','Орнамент':'Рисунок','Однотонная':'Рисунок','Водоотталкивающая':'Влаго-грязе-масло отталкивающая пропитка','Количество в упаковке, штук':'В наборе, шт.','Глажение':'Рекомендации по уходу','Отбелевание':'Рекомендации по уходу','Объем, мл':'Объём, мл','Можно мыть в посудомоечной машине':'Можно мыть в посудомоечной машине','Сушка':'Рекомендации по уходу','Тип упаковки':'Вид упаковки','Тип упаковки':'Индивидуальная упаковка','Цвет производителя':'Цвет','Размеры (ДхШ), см':'Размер (Длина × Ширина)','Ширина, в сантиметрах':'Размер (Длина × Ширина)','Толщина, в сантиметрах':'Толщина, мм','Ширина, в сантиметрах':'Размер, см','Длина, в сантиметрах':'Размер, см','Размеры (ДхШ), см':'Размер, см','Рисунок':'Рисунок','Основной цвет':'Цвет','Кружка-хамелеон':'Особенности','Ширина коробки, в сантиметрах':'Размер упаковки (Длина × Ширина × Высота)','Можно использовать в СВЧ-печи':'Можно использовать в СВЧ-печи','Размер изделия, в сантиметрах':'Размер (Длина × Ширина × Высота)','Крепление на горизонтальную плоскость':'Вид крепежа','Именная':'Выбор имени','Количество чашек или кружек, шт.':'Количество персон','С двойными стенками':'Особенности','Размеры, в сантиметрах':'Размер фото','Вес, в килограммах':'Вес брутто','Ложка в комплекте':'Дополнительный аксессуар','С крышкой':'Дополнительный аксессуар','С ситечком':'Составляющие набора','Тематика':'Тематика'}

countimage = 1

for celltocheck in column_l: #Идёт по колонке
    count = count + 1
    print(f'checking cell{count}')
    idcell = ws['A'+str(count)].value
    linkcell = ws['R'+str(count)].value #идёт по ссылке
    if linkcell == None:
        break
    if count % 3000 == 0:
        print('saving')
        wb.save(f'{count}checked.xlsx') 
        # break
    if linkcell != None:
        print(celltocheck.value)
        print(linkcell)
        if count >1:
            if idcell !=  ws['A'+str(count-1)].value and linkcell != ws['R'+str(count-1)].value: #если новая ссылка то открывает её
                print(ws['R'+str(count-1)])
                
                print(f'opening new link{linkcell}')
                
                r = session.get(linkcell)
                # r.html.render()
                time.sleep(1)
                matches = r.html.find('div.S3jLY') 

                dict = {}
                  
                savedvalues=[]
                checkedcolour = False    
                for match in matches: #смотрит все совпадения по тегу хтмл
                    print(match.text)
                    if 'Цветочно-растительная' in match.text:
                        continue
                    for i in neccessaryattributes:
                        if i in match.text:
                            
                                
                                    
                              
                            match2 = match.text
                            newlines = match2.splitlines()
                            savedvalues.append(newlines[1])
                            dict[i]=match.text.splitlines()[1] 
                            
                


                              
            if celltocheck.value in dictionary.keys():
                for k,v in dictionary.items():
                    if k == celltocheck.value:
                        try:
                            if k == 'Вес коробки, в килограммах':
                                changedweight = dict[v]
                                if ' г' in dict[v]:
                                    changedweight = changedweight.replace(' г','')
                                    number = int(changedweight) / 1000
                                    celltocheck.value = str(number)
                                    celltocheck.value = celltocheck.value.replace('.',',')
                                    print(number)
                                    print(f'I found value {celltocheck.value}')
                                elif 'кг' in dict[v]:
                                    celltocheck.value = dict[v]
                                    celltocheck.value = str(celltocheck.value).replace(' кг','')
                                    celltocheck.value = celltocheck.value.replace('.',',')
                                    print(f'I found value {celltocheck.value}')
                            elif k == 'Вес, в килограммах':
                                changedweight = dict[v]
                                if ' г' in dict[v]:
                                    changedweight = changedweight.replace(' г','')
                                    number = int(changedweight) / 1000
                                    celltocheck.value = str(number)
                                    celltocheck.value = celltocheck.value.replace('.',',')
                                    print(number)
                                    print(f'I found value {celltocheck.value}')
                                elif 'кг' in dict[v]:
                                    celltocheck.value = dict[v]
                                    celltocheck.value = str(celltocheck.value).replace(' кг','')
                                    celltocheck.value = celltocheck.value.replace('.',',')
                                    print(f'I found value {celltocheck.value}')
                            elif k == 'Длина коробки, в сантиметрах':
                                checkme = dict['Размер упаковки (Длина × Ширина × Высота)']
                                celltocheck.value = re.findall(r'\d+,?\d?',checkme)[0]
                                print(f'I found value {celltocheck.value}')
                                if dict[v] == None:
                                    print('NOT ROOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOUND')
                                    celltocheck.value = 'not found'
                            elif k == 'Ширина коробки, в сантиметрах':
                                checkme = dict['Размер упаковки (Длина × Ширина × Высота)']
                                celltocheck.value = re.findall(r'\d+,?\d?',checkme)[1]
                                print(f'I found value {celltocheck.value}')
                                if dict[v] == None:
                                    print('NOT ROOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOUND')
                                    celltocheck.value = 'not found'        
                            elif k == 'Длина, в сантиметрах':
                                if 'Размер (Длина × Ширина)' in dict:
                                    checkme = dict['Размер (Длина × Ширина)']
                                    celltocheck.value = re.findall(r'\d+,?\d?',checkme)[0]
                                    print(f'I found value {celltocheck.value}')
                                else:
                                    checkme = dict['Размер, см']
                                    celltocheck.value = re.findall(r'\d+,?\d?',checkme)[0]
                                    print(f'I found value {celltocheck.value}')
                            elif k == 'Ширина, в сантиметрах':
                                if 'Ширина, см' in dict:
                                    checkme = dict['Ширина, см']
                                    celltocheck.value = re.findall(r'\d+,?\d?',checkme)[0]
                                    print(f'I found value {celltocheck.value}')
                                else:
                                    checkme = dict['Размер, см']
                                    celltocheck.value = re.findall(r'\d+,?\d?',checkme)[0]
                                    print(f'I found value {celltocheck.value}')
                            elif k == 'Цвет':
                                celltocheck.value = dict[v].lower()
                                print(f'I found value {celltocheck.value}')
                            elif k == 'Цвет производителя (маркетинговый цвет)':
                                celltocheck.value = dict[v]
                                print(f'I found value {celltocheck.value}')
                            elif k == 'Материал':
                                if 'Материал' in dict:
                                    celltocheck.value = dict['Материал'].lower()
                                    print(f'I found value {celltocheck.value}')
                                else:
                                    celltocheck.value = dict['Состав ткани'].lower()
                                    print(f'I found value {celltocheck.value}')
                            elif k == 'Размер изделия, в сантиметрах':
                                celltocheck.value = dict[v]
                                print(f'I found value {celltocheck.value}')
                            elif k == 'Размеры, в сантиметрах':
                                celltocheck.value = dict[v]
                                print(f'I found value {celltocheck.value}')
                            elif k == 'Количество в упаковке, штук':
                                celltocheck.value = dict[v]
                                print(f'I found value {celltocheck.value}')
                            elif k == 'Коллекция':
                                celltocheck.value = dict[v]
                                print(f'I found value {celltocheck.value}')
                            elif k == 'Отжим':
                                celltocheck.value = dict[v]
                                print(f'I found value {celltocheck.value}')
                            elif k == 'Гипоаллергенная':
                                celltocheck.value = dict[v]
                                print(f'I found value {celltocheck.value}')
                            elif k == 'С крышкой':
                                celltocheck.value = dict[v]
                                print(f'I found value {celltocheck.value}')
                            elif k == 'Тип упаковки':
                                if 'Вид упаковки' in dict:
                                    celltocheck.value = dict['Вид упаковки']
                                    print(f'I found value {celltocheck.value}')
                                else:
                                    celltocheck.value = dict['Индивидуальная упаковка']
                                    print(f'I found value {celltocheck.value}')
                            elif k == 'Именная':
                                if 'Рисунок' in dict:
                                    celltocheck.value = 'Да'
                                    print(f'I found value {celltocheck.value}')
                                else:
                                    celltocheck.value = 'Нет'
                            elif k == 'Тип товара':
                                if dict['Объём, мл'] > 250:
                                    celltocheck.value = 'кружка'
                                else:
                                    celltocheck.value = 'чашка'
                                    print(f'I found value {celltocheck.value}')
                            elif k == 'Сушка':
                                celltocheck.value = dict[v]
                                print(f'I found value {celltocheck.value}')
                            elif k == 'Ложка в комплекте':
                                if 'Дополнительный аксессуар' in dict:
                                    if 'Ложка' in dict['Дополнительный аксессуар']:
                                        celltocheck.value = 'Да'
                                        print(f'I found value {celltocheck.value}')
                                    else:
                                        celltocheck.value = 'Нет'
                                else:
                                    celltocheck.value = 'Нет'
                            elif k == 'Особенности дизайна':
                                if 'Тематика праздника' in dict:
                                    celltocheck.value = dict['Тематика праздника']
                                    print(f'I found value {celltocheck.value}')
                                else:
                                    celltocheck.value = dict['Тематика']
                            elif k == 'С ситечком':
                                if 'Составляющие набора' in dict:
                                    if 'Сито' in dict['Дополнительный аксессуар']:
                                        celltocheck.value = 'Да'
                                        print(f'I found value {celltocheck.value}')
                                    else:
                                        celltocheck.value = 'Нет'
                                else:
                                    celltocheck.value = 'Нет'
                            elif k == 'С крышкой':
                                if 'Дополнительный аксессуар' in dict:
                                    if 'Крышка' in dict['Дополнительный аксессуар']:
                                        celltocheck.value = 'Да'
                                        print(f'I found value {celltocheck.value}')
                                    else:
                                        celltocheck.value = 'Нет'
                                else:
                                    celltocheck.value = 'Нет'
                            elif k == 'Подарочная упаковка':
                                print(celltocheck.value)
                                if 'Вид упаковки' in dict:
                                    if 'Подарочная коробка' in dict['Дополнительный аксессуар']:
                                        celltocheck.value = 'Да'
                                        #print(f'I found value {celltocheck.value}')
                                    else:
                                        celltocheck.value = 'Нет'
                                else:
                                    celltocheck.value = dict[v]
                                    print(celltocheck.value)
                            elif k == 'Высота изделия, в сантиметрах':
                                celltocheck.value = dict[v]
                                print(f'I found value {celltocheck.value}')
                            elif k == 'С двойными стенками':
                                if 'Особенности' in dict:
                                    if 'Двойные стенки' in dict['Дополнительный аксессуар']:
                                        celltocheck.value = 'Да'
                                        print(f'I found value {celltocheck.value}')
                                    else:
                                        celltocheck.value = 'Нет'
                                else:
                                    celltocheck.value = 'Нет'
                            elif k == 'Формат фото':
                                celltocheck.value = dict[v].lower()
                                print(f'I found value {celltocheck.value}')
                            elif k == 'Крепление на стену':
                                if 'Подвес' in dict[v]:
                                    celltocheck.value = 'Да'
                                    print(f'I found value {celltocheck.value}')
                                else:
                                    celltocheck.value = 'Нет'
                            elif k == 'Размеры (ДхШ), см':
                                if 'Размер, см' in dict:
                                    sizeit = dict['Размер, см']
                                    sizeit = sizeit.replace(' х ','x')
                                    sizeit = sizeit.replace(' см','')
                                    celltocheck.value = sizeit
                                    print(f'I found value {celltocheck.value}')
                                else:
                                    sizeit = dict['Размер (Длина × Ширина)']
                                    sizeit = sizeit.replace(' х ','x')
                                    sizeit = sizeit.replace(' см','')
                                    celltocheck.value = sizeit
                                    print(f'I found value {celltocheck.value}')
                            elif k == 'Модель товара':
                                checkname = ws['D'+str(count)].value
                               
                                if '"' in checkname:
                                  
                                    found = re.findall(r'"(.*?)"', checkname).group[0]
                                    celltocheck.value = found
                                elif '«' in checkname:
                                    
                                    found = re.findall(r'«(.*?)»', checkname).group[0]
                                    celltocheck.value = found
                                else:
                                    celltocheck.value = ''
                            elif k == 'Высота, в сантиметрах':
                                if 'Высота, см' in dict:
                                    checkme = dict['Высота, см']
                                    celltocheck.value = re.findall(r'\d+,?\d?',checkme)[0]
                                    print(f'I found value {celltocheck.value}')
                                else:
                                    checkme = dict['Размер, см']
                                    celltocheck.value = re.findall(r'\d+,?\d?',checkme)[0]
                                    print(f'I found value {celltocheck.value}')
                            elif k == 'Высота коробки, в сантиметрах':
                                print(dict[v])
                                heightarticle = re.findall(r'\d+,?\d?',dict[v])[2]
                                celltocheck.value = heightarticle
                                print(f'I found value {celltocheck.value}')
                            else:
                                celltocheck.value = dict[v]
                                print(f'I found value {celltocheck.value}')
                        except (KeyError,ValueError):
                            try:
                                if k == 'Длина коробки, в сантиметрах':
                                    checkme = dict['Размер упаковки (Длина × Ширина × Высота)']
                                    celltocheck.value = re.findall(r'\d+,?\d?',checkme)[0]
                                elif k == 'Ширина коробки, в сантиметрах':
                                    checkme = dict['Размер упаковки (Длина × Ширина × Высота)']
                                    celltocheck.value = re.findall(r'\d+,?\d?',checkme)[1]
                                elif k == 'Ширина, в сантиметрах':
                                    checkme = dict['Размер упаковки (Длина × Ширина × Высота)']
                                    celltocheck.value = re.findall(r'\d+,?\d?',checkme)[1]
                                elif k == 'Однотонная':
                                    celltocheck.value = dict[v]
                                    print(f'I found value {celltocheck.value}')
                                elif k == 'Орнамент':
                                    celltocheck.value = dict[v]
                                    print(f'I found value {celltocheck.value}')
                                else:
                                    celltocheck.value = ''
                                    print('ERRRRRRROR')
                            except (KeyError,ValueError):
                                celltocheck.value = ''
           
               
            else:
                celltocheck.value = ''
                print('no value')
                
print(dict)    

wb.save('xxx')    
print(datetime.now() - startTime)                 
