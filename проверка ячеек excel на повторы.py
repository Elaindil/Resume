import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
startTime = datetime.now()


wb = load_workbook('test.xlsx')
print('loaded workbook')
ws = wb['Лист1']

column_b = ws['B']

column_r = ws['R']
ft = Font(color="FF0000")
count = 1
linktocheck = ''

foundlinka = False

differencefound = False
count = 1
for columnb in column_b:
    
    if columnb.value == None:
        break 
    
    if count >1:
        if columnb.value !=  ws['B'+str(count-1)].value and columnb.value.lower() == ws['B'+str(count-1)].value.lower(): #если новая ссылка то открывает её
            ws['C'+str(count)].value = count
            ws['C'+str(count-1)].value = count
            # if ws['R'+str(count-1)].value ==  ws['R'+str(count)].value:
                # print('error')
                # print(f'{count}')
                # colourme = ws['R'+str(count)]
                # colourme.font = ft

    count = count + 1


wb.save('checkedlinks.xlsx')  
print(datetime.now() - startTime)   
