from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

startTime = datetime.now()
options = Options()


wb = load_workbook('xxxx.xlsx')
print('loaded workbook')
ws = wb['WorkSheet']

column_A = ws['A']

column_l = ws['L']

count = 1
linktocheck = ''




for columna in column_A:
    if columna.value == None:
        break  
    
    if count >1:
        if columna.value !=  ws['A'+str(count-1)].value and ws['V'+str(count)].value == 'wrong': #если новая ссылка то открывает её
            searchstring = ws['I'+str(count)].value

            driver = webdriver.Firefox(options=options)
            driver.get("https://www.google.com/")
            x = driver.find_element(By.CLASS_NAME,'Gdd5U').click()
            time.sleep(1)
            i = driver.find_element(By.CLASS_NAME,'cB9M7').send_keys(searchstring)
            time.sleep(1)
            x = driver.find_element(By.CLASS_NAME,'Qwbd3').click()
            time.sleep(1)
            listi= driver.find_elements(By.CLASS_NAME,'GZrdsf.lXbkTc')


            for i in listi:
                if 'sima-land' in i.get_attribute('href'):
                    print(i.get_attribute('href'))
                    ws['S'+str(count)].value = i.get_attribute('href')
                    driver.close() 
                    break

            

      

    if count % 3000 == 0:
        wb.save(f'{count}addedlinks.xlsx') 
    count = count + 1
    print(count)
wb.save('addedlinks.xlsx')  
driver.close() 
                    

print(datetime.now() - startTime)   

