import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
startTime = datetime.now()


wb = load_workbook('xx.xlsx')
print('loaded workbook')
ws = wb['Лист1']

column_b = ws['B']
column_c = ws['C']
column_e = ws['E']
ft = Font(color="FF0000", size=12)

redFill = PatternFill(start_color='FFFF0000',end_color='FFFF0000',fill_type='solid')



count = 1
for columnb in column_b:
    checkwords = ws['D'+str(count)].value
    if checkwords == None:
        break
    
    checkwords = checkwords.split()

    for word in checkwords:
        if word.lower() in columnb.value.lower():
            columnb.fill = redFill
            print(f'{word} in {columnb.value} for brand {checkwords}')
            break

                
                
                
               
            
    count = count + 1

count = 1
for columnc in column_c:
    checkwords = ws['D'+str(count)].value
    if checkwords == None:
        break
    print(count)
    checkwords = checkwords.split()
    print(checkwords)
    for word in checkwords:
        if word.lower() in columnc.value.lower():
            columnc.fill = redFill
            print(f'{word} in {columnc.value} for brand {checkwords}')
            break

                
            
    count = count + 1
    



wb.save('xxxx.xlsx')  
print(datetime.now() - startTime)   
