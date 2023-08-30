from requests_html import HTMLSession, HTML
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import time
from datetime import datetime
import re
import os


startTime = datetime.now()

wb = load_workbook('test.xlsx')
print('opened workbook')
ws = wb['WorkSheet']

column_r = ws['R']

column_a = ws['A']

count = 1
session = HTMLSession()
# for columng in column_g:
    # if columng.value != None:
        # ws['R'+str(count)].value = 'https://www.sima-land.ru/'+str(columng.value)
        # print('Adding new link')
    # count = count + 1
# print('saving new workbook')    
# wb.save('testwithlinks.xlsm')
# print('saved new workbook') 
countimage = 0

for celltocheck in column_r: #Идёт по колонке
    count = count + 1
    print(f'checking cell{count}')
    linkcell = ws['R'+str(count)].value #идёт по ссылке
    if linkcell == None:
        break
    if linkcell != None:
        print(celltocheck.value)
        print(linkcell)
        if count >1:
            if linkcell !=  ws['R'+str(count-1)].value: #если новая ссылка то открывает её
                countimage = 0
                print(ws['R'+str(count-1)])
                
                print(f'opening new link{linkcell}')
                
                r = session.get(linkcell)
                # r.html.render()
                time.sleep(8)
                matches = r.html.find('img.gyoE4') 
                for match in matches:
                    time.sleep(1)
                    ilink = match.attrs["src"].replace('140.jpg','700.jpg')
                    r2 = session.get(ilink)
                    idname= ws['A'+str(count)].value
                    with open(f'{idname}b{countimage}.jpg', 'wb') as w: 
                        w.write(r2.content)
                        countimage = countimage + 1
                                # print(ilink)



for celltocheck in column_r: #Идёт по колонке ПЕРВОЕ ФОТО
    count = count + 1
    print(f'checking cell{count}')
    linkcell = ws['R'+str(count)].value
    idcell = ws['A'+str(count)].value#идёт по ссылке
    imagelink = ws['I'+str(count+1)].value
    print(ws['I'+str(count-1)])
    if imagelink == None:
        break
    if imagelink != None:
        print(celltocheck.value)
        print(linkcell)
        
        if count >1:
            if imagelink !=  ws['I'+str(count-1)].value and idcell !=  ws['A'+str(count-1)].value: #если новая ссылка то открывает её
                countimage = 0
                time.sleep(1)
                
                print(f'opening new link{linkcell}')
                idname= ws['A'+str(count)].value
                r2 = session.get(imagelink)
                with open(f'{idname}b{countimage}.jpg', 'wb') as w: 
                        w.write(r2.content)


# session = HTMLSession()

              
# r = session.get("https://cdn3.static1-sima-land.com/items/4204827/0/700.jpg?v=1585658293")
# r = session.get("https://www.sima-land.ru/6959015/blyudo-dlya-podachi-sapfir-21-12-5-cm-cvet-siniy/")
# r.html.render()
time.sleep(3)
# matches = r.html.find('img.gyoE4') 
# for match in matches:
    # count = count +1
    # ilink = match.attrs["src"].replace('140.jpg','700.jpg')
    # r2 = session.get(ilink)
    
    # with open(f'{count}.jpg', 'wb') as w: 
        # w.write(r2.content)
    # print(ilink)


          
print(datetime.now() - startTime)                 
