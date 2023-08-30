# -*- coding: utf-8 -*-

################################################################################
## Form generated from reading UI file 'untitled.ui'
##
## Created by: Qt User Interface Compiler version 6.5.0
##
## WARNING! All changes made in this file will be lost when recompiling UI file!
################################################################################

from PySide6.QtCore import (QCoreApplication, QDate, QDateTime, QLocale,
    QMetaObject, QObject, QPoint, QRect,
    QSize, QTime, QUrl, Qt,)
from PySide6.QtGui import (QBrush, QColor, QConicalGradient, QCursor,
    QFont, QFontDatabase, QGradient, QIcon,
    QImage, QKeySequence, QLinearGradient, QPainter,
    QPalette, QPixmap, QRadialGradient, QTransform,QPixmap,)
from PySide6.QtWidgets import (QApplication, QGraphicsView, QMainWindow, QMenuBar,
    QPushButton, QSizePolicy, QStatusBar, QWidget,QFileDialog,QGraphicsPixmapItem,QLabel)
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from pathlib import Path
import glob
import requests
from requests_html import HTMLSession, HTML
ft = Font(color="FF0000")
# count = 1

canGo = True
count = 18113
wb = ''
testimage ='1.jpg'
url = ''
url2 = ''
ilink = ''
session = HTMLSession()
def loopLine():
    global count
    global url
    global url2
    global ilink
    print('looping')
    for i in range(count,99999):
        
        lineCheck = ws['A'+str(count)].value
        linkcell = ws['R'+str(count)].value
        origImage = ws['I'+str(count+1)].value
        global canGo
        if lineCheck == None:
            print('error')
            canGo = False
            wb.save('checkedlinks.xlsx') 
            break
           
        if count >1 and lineCheck !=  ws['A'+str(count-1)].value and canGo == True:
            canGo = False
            print('new')
            # session = HTMLSession()
            r = session.get(linkcell)
            matches = r.html.find('img.gyoE4')
            for match in matches:
                # time.sleep(1)
                ilink = match.attrs["src"].replace('140.jpg','700.jpg')
                # r2 = session.get(ilink)
                break
            break
        # session = HTMLSession()
        # r = session.get(linkcell)
        # matches = r.html.find('img.gyoE4')
        # for match in matches:
            # time.sleep(1)
            # ilink = match.attrs["src"].replace('140.jpg','700.jpg')
            # r2 = session.get(ilink)
            # break
        print(lineCheck)
        print(count)
        url2 = ilink
        # url =  ws['I'+str(count)].value
        url =  origImage
        count = count + 1
        print(url2)

def saveFileNow():
    wb.save(f'checkedlinks{count}.xlsx') 
    print('saved')

def openFile():
    fname = QFileDialog.getOpenFileName()
    fileToCheck = fname[0]
    print (fileToCheck)
    global wb
    wb = load_workbook(fileToCheck)
    print('loaded workbook')
    global ws
    ws = wb.active
    column_a = ws['A']
    loopLine()
   


 
def differentImage():
    global count
    global ws
    count +=1
    colourme = ws['D'+str(count-1)]
    colourme.font = ft
    global canGo
    canGo = True
    print(count)
    # global testimage
    currentline = ws['A'+str(count)].value
    # testimage = f"{currentline}.jpg"
    # testimage = data
    ws['W'+str(count-1)].value = 'wrong'
    loopLine()

def sameImage():
    global canGo
    global ws
    canGo = True
    global count
    global testimage
    currentline = ws['A'+str(count)].value
    testimage = f"{currentline}b1.jpg"
    count +=1
    print(testimage)
    loopLine()


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        if not MainWindow.objectName():
            MainWindow.setObjectName(u"MainWindow")
        MainWindow.resize(1200, 800)
        self.centralwidget = QWidget(MainWindow)
        self.centralwidget.setObjectName(u"centralwidget")
        self.chooseFileButton = QPushButton(self.centralwidget)
        self.chooseFileButton.setObjectName(u"chooseFileButton")
        self.chooseFileButton.setGeometry(QRect(80, 50, 75, 24))
        self.saveNowButton = QPushButton(self.centralwidget)
        self.saveNowButton.setObjectName(u"Save Now")
        self.saveNowButton.setGeometry(QRect(600, 50, 75, 24))
        self.differentButton = QPushButton(self.centralwidget)
        self.differentButton.setObjectName(u"differentButton")
        self.differentButton.setGeometry(QRect(120, 740, 75, 24))
        self.sameButton = QPushButton(self.centralwidget)
        self.sameButton.setObjectName(u"sameButton")
        self.sameButton.setGeometry(QRect(520, 730, 75, 24))
        self.graphicsView = QLabel(self.centralwidget)
        self.graphicsView.setObjectName(u"graphicsView")
        self.graphicsView.setGeometry(QRect(70, 140, 456, 492))
        self.graphicsView_2 = QLabel(self.centralwidget)
        self.graphicsView_2.setObjectName(u"graphicsView_2")
        self.graphicsView_2.setGeometry(QRect(470, 150, 456, 492))
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QMenuBar(MainWindow)
        self.menubar.setObjectName(u"menubar")
        self.menubar.setGeometry(QRect(0, 0, 800, 22))
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QStatusBar(MainWindow)
        self.statusbar.setObjectName(u"statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)

        QMetaObject.connectSlotsByName(MainWindow)
        
        self.saveNowButton.clicked.connect(lambda:saveFileNow())
        
        self.chooseFileButton.clicked.connect(lambda:openFile())
        self.chooseFileButton.clicked.connect(self.setImageOriginal)
        self.chooseFileButton.clicked.connect(self.setImageFound)
        
        
        self.differentButton.clicked.connect(lambda:differentImage())
        self.differentButton.clicked.connect(self.setImageOriginal)
        self.differentButton.clicked.connect(self.setImageFound)
        self.differentButton.setShortcut('Q')
        self.sameButton.clicked.connect(lambda:sameImage())

        self.sameButton.clicked.connect(self.setImageOriginal)
        self.sameButton.clicked.connect(self.setImageFound)
        self.sameButton.setShortcut('E')
   

        
        # pix2 = QPixmap('1.jpg')
        # pix2 = pix2.scaledToWidth(252)
        # pix2 = pix2.scaledToHeight(252)
        # self.graphicsView_2.setPixmap(pix2) 
        
        
    def setImageOriginal(self):
            image = QImage()
            image.loadFromData(requests.get(url).content)
            pix = QPixmap(image)
            pix = pix.scaledToWidth(552)
            pix = pix.scaledToHeight(552)
            self.graphicsView.setPixmap(pix)
            
    def setImageFound(self):
            # pix2 = QPixmap(testimage)
            image2 = QImage()
            image2.loadFromData(requests.get(ilink).content)
            pix2 = QPixmap(image2)
            pix2 = pix2.scaledToWidth(552)
            pix2 = pix2.scaledToHeight(552)
            self.graphicsView_2.setPixmap(pix2) 
            # pix = QPixmap(testimage)
            # pix = pix.scaledToWidth(252)
            # pix = pix.scaledToHeight(252)
        # item = QGraphicsPixmapItem(pix)
        # scene = QGraphicsScence(self)
        # scene.addItem(item)
        # self.ui.graphicsView.setScene(scene)
                
        
    # setupUi

    def retranslateUi(self, MainWindow):
        MainWindow.setWindowTitle(QCoreApplication.translate("MainWindow", u"MainWindow", None))
        self.chooseFileButton.setText(QCoreApplication.translate("MainWindow", u"Choose file", None))
        self.saveNowButton.setText(QCoreApplication.translate("MainWindow", u"Save now", None))
        self.differentButton.setText(QCoreApplication.translate("MainWindow", u"Different", None))
        self.sameButton.setText(QCoreApplication.translate("MainWindow", u"Same", None))
    # retranslateUi



import sys
if __name__ == "__main__":
    app = QApplication(sys.argv)
    ex = Ui_MainWindow()
    w = QMainWindow()
    ex.setupUi(w)
    w.show()
    sys.exit(app.exec())
    